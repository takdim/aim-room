import datetime as dt
import os

import io

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models.holiday import Holiday
from app.models.pakta_template import PaktaTemplate
from app.models.reference import Building, Course
from app.models.class_schedule import ClassSchedule, class_schedule_lecturers
from app.models.lecturer import Lecturer
from app.models.room import Room
from app.models.room_booking import RoomBooking
from app.models.user import User, ROLES, APPROVAL_ROLES
from app.models.semester import Semester
from app.routes.auth import login_required, role_required

dashboard_bp = Blueprint("dashboard", __name__, url_prefix="/dashboard")


def _allowed_file(filename: str) -> bool:
    allowed = current_app.config.get("ALLOWED_EXTENSIONS", {"pdf", "jpg", "jpeg", "png"})
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


def _save_upload(file, subfolder: str) -> str | None:
    if not file or not file.filename:
        return None
    if not _allowed_file(file.filename):
        return None
    upload_root = current_app.config["UPLOAD_FOLDER"]
    dest_dir = os.path.join(upload_root, subfolder)
    os.makedirs(dest_dir, exist_ok=True)
    filename = secure_filename(file.filename)
    unique_name = f"{int(dt.datetime.utcnow().timestamp())}_{filename}"
    file.save(os.path.join(dest_dir, unique_name))
    return os.path.join(subfolder, unique_name)


def _get_active_semester():
    """Get semester from session, fallback to is_active=True for schedules."""
    semester_id = session.get("active_semester_id")
    if semester_id:
        semester = Semester.query.get(int(semester_id))
        if semester:
            return semester
    return Semester.query.filter_by(is_active=True).order_by(Semester.id.desc()).first()


def _get_selected_semester():
    """Get semester from session only. Return None if not selected."""
    semester_id = session.get("active_semester_id")
    if semester_id:
        semester = Semester.query.get(int(semester_id))
        if semester:
            return semester
    return None


@dashboard_bp.route("/admin", methods=["GET", "POST"])
@login_required
@role_required("admin")
def admin_home():
    if request.method == "POST":
        action = request.form.get("action", "create")

        # ── DELETE ──────────────────────────────────────────────
        if action == "delete":
            user_id = request.form.get("user_id", "").strip()
            user = User.query.get(int(user_id)) if user_id else None
            if user and user.role != "admin":
                db.session.delete(user)
                db.session.commit()
                flash("Akun berhasil dihapus.", "info")
            else:
                flash("Akun tidak ditemukan atau tidak dapat dihapus.", "error")
            return redirect(url_for("dashboard.admin_home"))

        # ── UPDATE ──────────────────────────────────────────────
        if action == "update":
            user_id = request.form.get("user_id", "").strip()
            user = User.query.get(int(user_id)) if user_id else None
            if not user or user.role == "admin":
                flash("Akun tidak ditemukan.", "error")
                return redirect(url_for("dashboard.admin_home"))

            full_name = request.form.get("full_name", "").strip()
            email = request.form.get("email", "").strip().lower()
            role = request.form.get("role", user.role).strip()
            new_password = request.form.get("password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()

            if not full_name or not email:
                flash("Nama dan email wajib diisi.", "error")
                return redirect(url_for("dashboard.admin_home"))

            if role not in ROLES or role == "admin":
                role = "staff"

            conflict = User.query.filter(User.email == email, User.id != user.id).first()
            if conflict:
                flash("Email sudah digunakan akun lain.", "error")
                return redirect(url_for("dashboard.admin_home"))

            if new_password:
                if new_password != confirm_password:
                    flash("Konfirmasi password tidak cocok.", "error")
                    return redirect(url_for("dashboard.admin_home"))
                if len(new_password) < 6:
                    flash("Password minimal 6 karakter.", "error")
                    return redirect(url_for("dashboard.admin_home"))
                user.set_password(new_password)

            user.full_name = full_name
            user.email = email
            user.role = role
            db.session.commit()
            flash("Akun berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.admin_home"))

        # ── CREATE (default) ────────────────────────────────────
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        role = request.form.get("role", "staff").strip()

        if not full_name or not email or not password:
            flash("Nama, email, dan password wajib diisi.", "error")
            return redirect(url_for("dashboard.admin_home"))

        if len(password) < 6:
            flash("Password minimal 6 karakter.", "error")
            return redirect(url_for("dashboard.admin_home"))

        if password != confirm_password:
            flash("Konfirmasi password tidak cocok.", "error")
            return redirect(url_for("dashboard.admin_home"))

        if role not in ROLES or role == "admin":
            role = "staff"

        if User.query.filter_by(email=email).first():
            flash("Email sudah digunakan.", "error")
            return redirect(url_for("dashboard.admin_home"))

        new_user = User(full_name=full_name, email=email, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash(f"Akun {role} berhasil dibuat.", "info")
        return redirect(url_for("dashboard.admin_home"))

    all_users = User.query.filter(User.role != "admin").order_by(User.id.desc()).all()
    total_users = User.query.count()
    admin_count = User.query.filter_by(role="admin").count()
    return render_template(
        "dashboard/admin.html",
        full_name=session.get("full_name"),
        all_users=all_users,
        total_users=total_users,
        admin_count=admin_count,
        roles=[r for r in ROLES if r != "admin"],
    )


@dashboard_bp.get("/staff")
@login_required
@role_required("staff")
def staff_home():
    semesters = Semester.query.order_by(Semester.id.desc()).all()
    active_semester = _get_active_semester()
    
    # Filter stats berdasarkan semester yang dipilih
    if active_semester:
        # Courses yang ada di schedule semester ini
        total_courses = db.session.query(ClassSchedule.course_id).filter(
            ClassSchedule.semester_id == active_semester.id
        ).distinct().count()
        
        # Lecturers yang mengajar di semester ini
        total_lecturers = (
            db.session.query(class_schedule_lecturers.c.lecturer_id)
            .join(ClassSchedule, ClassSchedule.id == class_schedule_lecturers.c.schedule_id)
            .filter(ClassSchedule.semester_id == active_semester.id)
            .distinct()
            .count()
        )
        
        # Rooms yang digunakan di semester ini
        total_rooms = db.session.query(ClassSchedule.room_id).filter(
            ClassSchedule.semester_id == active_semester.id
        ).distinct().count()
    else:
        # Jika tidak ada semester yang dipilih, tampilkan 0
        total_courses = 0
        total_lecturers = 0
        total_rooms = 0
    
    return render_template(
        "dashboard/staff.html",
        full_name=session.get("full_name"),
        total_courses=total_courses,
        total_lecturers=total_lecturers,
        total_rooms=total_rooms,
        semesters=semesters,
        active_semester=active_semester,
    )


@dashboard_bp.route("/staff/courses", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_courses():
    if request.method == "POST":
        action = request.form.get("action", "add")
        course_id = request.form.get("course_id", "").strip()
        course_name = request.form.get("course_name", "").strip()
        course_code = request.form.get("course_code", "").strip().upper()

        if action == "delete":
            if course_id:
                Course.query.filter_by(id=int(course_id)).delete()
                db.session.commit()
                flash("Kelas berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_courses"))

        if not course_name:
            flash("Nama kelas wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_courses"))

        if action == "update":
            course = Course.query.get(int(course_id)) if course_id else None
            if not course:
                flash("Kelas tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_courses"))
            if course_code and Course.query.filter(Course.course_code == course_code, Course.id != course.id).first():
                flash("Kode kelas sudah digunakan.", "error")
                return redirect(url_for("dashboard.staff_courses"))
            course.course_name = course_name
            course.course_code = course_code or None
            db.session.commit()
            flash("Kelas berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_courses"))

        if course_code and Course.query.filter_by(course_code=course_code).first():
            flash("Kode kelas sudah digunakan.", "error")
            return redirect(url_for("dashboard.staff_courses"))
        db.session.add(Course(course_name=course_name, course_code=course_code or None))
        db.session.commit()
        flash("Kelas berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_courses"))

    q = request.args.get("q", "").strip()
    active_semester = _get_active_semester()
    
    # Tampilkan SEMUA courses tanpa filter semester
    query = Course.query
    if q:
        query = query.filter(
            or_(
                Course.course_name.ilike(f"%{q}%"),
                Course.course_code.ilike(f"%{q}%"),
            )
        )
    courses = query.order_by(Course.id.desc()).limit(50).all()
    return render_template(
        "dashboard/staff_courses.html",
        full_name=session.get("full_name"),
        courses=courses,
        q=q,
        active_semester=active_semester,
    )


@dashboard_bp.route("/staff/lecturers", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_lecturers():
    if request.method == "POST":
        action = request.form.get("action", "add")
        lecturer_id = request.form.get("lecturer_id", "").strip()
        lecturer_name = request.form.get("lecturer_name", "").strip()
        nidn = request.form.get("nidn", "").strip()

        if action == "delete":
            if lecturer_id:
                Lecturer.query.filter_by(id=int(lecturer_id)).delete()
                db.session.commit()
                flash("Pengajar berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_lecturers"))

        if not lecturer_name:
            flash("Nama pengajar wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_lecturers"))

        if action == "update":
            lecturer = Lecturer.query.get(int(lecturer_id)) if lecturer_id else None
            if not lecturer:
                flash("Pengajar tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_lecturers"))
            lecturer.lecturer_name = lecturer_name
            lecturer.nidn = nidn or None
            db.session.commit()
            flash("Pengajar berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_lecturers"))

        db.session.add(Lecturer(lecturer_name=lecturer_name, nidn=nidn or None))
        db.session.commit()
        flash("Pengajar berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_lecturers"))

    q = request.args.get("q", "").strip()
    active_semester = _get_active_semester()
    
    # Tampilkan SEMUA lecturers tanpa filter semester
    query = Lecturer.query
    if q:
        query = query.filter(
            or_(
                Lecturer.lecturer_name.ilike(f"%{q}%"),
                Lecturer.nidn.ilike(f"%{q}%"),
            )
        )
    lecturers = query.order_by(Lecturer.id.desc()).limit(50).all()
    return render_template(
        "dashboard/staff_lecturers.html",
        full_name=session.get("full_name"),
        lecturers=lecturers,
        q=q,
        active_semester=active_semester,
    )


@dashboard_bp.route("/staff/rooms", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_rooms():
    if request.method == "POST":
        action = request.form.get("action", "add")
        room_id = request.form.get("room_id", "").strip()
        room_code = request.form.get("room_code", "").strip()
        room_name = request.form.get("room_name", "").strip()
        building_id = request.form.get("building_id", "").strip()
        floor = request.form.get("floor", "").strip()
        capacity = request.form.get("capacity", "").strip()
        room_type = request.form.get("room_type", "").strip()

        if action == "delete":
            if room_id:
                Room.query.filter_by(id=int(room_id)).delete()
                db.session.commit()
                flash("Ruangan berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_rooms"))

        if not room_name:
            flash("Nama ruangan wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_rooms"))

        building_id_value = int(building_id) if building_id else None
        if building_id_value and not Building.query.get(building_id_value):
            flash("ID gedung tidak ditemukan, nilai dikosongkan.", "error")
            building_id_value = None

        if action == "update":
            room = Room.query.get(int(room_id)) if room_id else None
            if not room:
                flash("Ruangan tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_rooms"))
            room.room_code = room_code or None
            room.room_name = room_name
            room.building_id = building_id_value
            room.floor = int(floor) if floor else None
            room.capacity = int(capacity) if capacity else None
            room.room_type = room_type or None
            db.session.commit()
            flash("Ruangan berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_rooms"))

        room = Room(
            room_code=room_code or None,
            room_name=room_name,
            building_id=building_id_value,
            floor=int(floor) if floor else None,
            capacity=int(capacity) if capacity else None,
            room_type=room_type or None,
        )
        db.session.add(room)
        db.session.commit()
        flash("Ruangan berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_rooms"))

    q = request.args.get("q", "").strip()
    active_semester = _get_active_semester()
    
    # Tampilkan SEMUA rooms tanpa filter semester
    query = Room.query
    if q:
        query = query.filter(
            or_(
                Room.room_name.ilike(f"%{q}%"),
                Room.room_code.ilike(f"%{q}%"),
                Room.room_type.ilike(f"%{q}%"),
            )
        )
    rooms = query.order_by(Room.id.desc()).limit(50).all()
    buildings = Building.query.order_by(Building.id.asc()).all()
    return render_template(
        "dashboard/staff_rooms.html",
        full_name=session.get("full_name"),
        rooms=rooms,
        buildings=buildings,
        q=q,
        active_semester=active_semester,
    )


@dashboard_bp.route("/staff/buildings", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_buildings():
    if request.method == "POST":
        action = request.form.get("action", "add")
        building_id = request.form.get("building_id", "").strip()
        building_id_input = request.form.get("building_id_input", "").strip()
        building_name = request.form.get("building_name", "").strip()

        if action == "delete":
            if building_id:
                Building.query.filter_by(id=int(building_id)).delete()
                db.session.commit()
                flash("Gedung berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_buildings"))

        if not building_name:
            flash("Nama gedung wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_buildings"))

        if action == "update":
            building = Building.query.get(int(building_id)) if building_id else None
            if not building:
                flash("Gedung tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_buildings"))
            building.building_name = building_name
            db.session.commit()
            flash("Gedung berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_buildings"))

        building = Building(building_name=building_name)
        if building_id_input:
            building.id = int(building_id_input)
        db.session.add(building)
        try:
            db.session.commit()
            flash("Gedung berhasil ditambahkan.", "info")
        except IntegrityError:
            db.session.rollback()
            flash("ID gedung sudah digunakan.", "error")
        return redirect(url_for("dashboard.staff_buildings"))

    buildings = Building.query.order_by(Building.id.desc()).limit(50).all()
    return render_template(
        "dashboard/staff_buildings.html",
        full_name=session.get("full_name"),
        buildings=buildings,
        active_semester=_get_active_semester(),
    )


@dashboard_bp.route("/staff/bookings", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_bookings():
    if request.method == "POST":
        action = request.form.get("action", "")
        booking_id = request.form.get("booking_id", "").strip()
        booking = RoomBooking.query.get(int(booking_id)) if booking_id else None

        if action == "delete" and booking:
            db.session.delete(booking)
            db.session.commit()
            flash("Peminjaman berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_bookings"))

        if action == "forward" and booking:
            # Staff checks done → forward to Kasubag
            booking.is_fib_student = request.form.get("is_fib_student") == "1"
            booking.is_bem_verified = request.form.get("is_bem_verified") == "1"
            booking.staff_note = request.form.get("staff_note", "").strip() or None
            booking.status = "Menunggu Kasubag"
            db.session.commit()
            flash("Pengajuan diteruskan ke Kasubag.", "info")
            return redirect(url_for("dashboard.staff_bookings"))

        if action == "reject" and booking:
            booking.status = "Ditolak"
            booking.rejection_note = request.form.get("rejection_note", "").strip() or None
            booking.rejected_by = "staff"
            db.session.commit()
            flash("Peminjaman ditolak.", "info")
            return redirect(url_for("dashboard.staff_bookings"))

    booking_rows = (
        db.session.query(RoomBooking, Room)
        .join(Room, RoomBooking.room_id == Room.id)
        .order_by(RoomBooking.booking_date.asc(), RoomBooking.id.desc())
        .limit(200)
        .all()
    )

    active_semester = _get_active_semester()
    return render_template(
        "dashboard/staff_bookings.html",
        full_name=session.get("full_name"),
        booking_rows=booking_rows,
        active_semester=active_semester,
    )


@dashboard_bp.get("/staff/bookings/<int:booking_id>")
@login_required
@role_required("staff", "admin")
def staff_booking_detail(booking_id: int):
    booking = RoomBooking.query.get_or_404(booking_id)
    return render_template(
        "dashboard/staff_booking_detail.html",
        full_name=session.get("full_name"),
        booking=booking,
        active_semester=_get_active_semester(),
    )


@dashboard_bp.get("/staff/bookings/file/<int:booking_id>/<field>")
@login_required
@role_required("staff", "admin")
def staff_serve_file(booking_id: int, field: str):
    if field not in ("pakta_integritas_path", "surat_permohonan_path"):
        from flask import abort
        abort(404)
    booking = RoomBooking.query.get_or_404(booking_id)
    file_path = getattr(booking, field, None)
    if not file_path:
        from flask import abort
        abort(404)
    upload_folder = current_app.config["UPLOAD_FOLDER"]
    dir_path = os.path.dirname(os.path.join(upload_folder, file_path))
    filename = os.path.basename(file_path)
    return send_from_directory(dir_path, filename, as_attachment=False)



@dashboard_bp.route("/staff/schedules", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_schedules():
    if request.method == "POST":
        action = request.form.get("action", "add")
        schedule_id = request.form.get("schedule_id", "").strip()
        course_id = request.form.get("course_id", "").strip()
        lecturer_ids = request.form.getlist("lecturer_ids")
        class_name = request.form.get("class_name", "").strip()
        room_id = request.form.get("room_id", "").strip()
        day_name = request.form.get("day_name", "").strip()
        start_time = request.form.get("start_time", "").strip()
        end_time = request.form.get("end_time", "").strip()
        semester_id = request.form.get("semester_id", "").strip()

        if action == "delete":
            if schedule_id:
                sched = ClassSchedule.query.get(int(schedule_id))
                if sched:
                    sched.lecturers = []
                    db.session.flush()
                    db.session.delete(sched)
                db.session.commit()
                flash("Jadwal berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_schedules"))

        if not (course_id and lecturer_ids and room_id and day_name and start_time and end_time and semester_id):
            flash("Semua field jadwal wajib diisi (minimal 1 pengajar).", "error")
            return redirect(url_for("dashboard.staff_schedules"))

        lecturers_list = Lecturer.query.filter(Lecturer.id.in_([int(lid) for lid in lecturer_ids])).all()

        if action == "update":
            schedule = ClassSchedule.query.get(int(schedule_id)) if schedule_id else None
            if not schedule:
                flash("Jadwal tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_schedules"))
            schedule.course_id = int(course_id)
            schedule.class_name = class_name or None
            schedule.room_id = int(room_id)
            schedule.day_name = day_name
            schedule.start_time = start_time
            schedule.end_time = end_time
            schedule.semester_id = int(semester_id)
            schedule.lecturers = lecturers_list
            db.session.commit()
            flash("Jadwal berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_schedules"))

        schedule = ClassSchedule(
            course_id=int(course_id),
            class_name=class_name or None,
            room_id=int(room_id),
            day_name=day_name,
            start_time=start_time,
            end_time=end_time,
            semester_id=int(semester_id),
        )
        schedule.lecturers = lecturers_list
        db.session.add(schedule)
        db.session.commit()
        flash("Jadwal berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_schedules"))

    q   = request.args.get("q", "").strip()
    day = request.args.get("day", "").strip()
    active_semester = _get_active_semester()
    query = (
        db.session.query(ClassSchedule, Course, Room)
        .join(Course, ClassSchedule.course_id == Course.id)
        .join(Room, ClassSchedule.room_id == Room.id)
    )
    if active_semester:
        query = query.filter(ClassSchedule.semester_id == active_semester.id)
    else:
        # Jika tidak ada semester yang dipilih, tidak tampilkan jadwal apapun
        query = query.filter(False)

    if day:
        query = query.filter(ClassSchedule.day_name == day)

    if q:
        query = query.filter(
            or_(
                Course.course_name.ilike(f"%{q}%"),
                Course.course_code.ilike(f"%{q}%"),
                ClassSchedule.class_name.ilike(f"%{q}%"),
                Room.room_name.ilike(f"%{q}%"),
                ClassSchedule.day_name.ilike(f"%{q}%"),
                ClassSchedule.lecturers.any(Lecturer.lecturer_name.ilike(f"%{q}%")),
            )
        )
    page       = request.args.get("page", 1, type=int)
    per_page   = 25
    pagination = (
        query
        .order_by(ClassSchedule.day_name.asc(), ClassSchedule.start_time.asc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    courses = Course.query.order_by(Course.course_name.asc()).all()
    lecturers = Lecturer.query.order_by(Lecturer.lecturer_name.asc()).all()
    rooms = Room.query.order_by(Room.room_name.asc()).all()
    semesters = Semester.query.order_by(Semester.id.desc()).all()

    return render_template(
        "dashboard/staff_schedules.html",
        full_name=session.get("full_name"),
        schedules=pagination.items,
        pagination=pagination,
        courses=courses,
        lecturers=lecturers,
        rooms=rooms,
        semesters=semesters,
        q=q,
        day=day,
        active_semester=active_semester,
    )


@dashboard_bp.route("/staff/semesters", methods=["GET", "POST"])
@login_required
@role_required("staff")
def staff_semesters():
    if request.method == "POST":
        action = request.form.get("action", "add")
        semester_id = request.form.get("semester_id", "").strip()
        name = request.form.get("name", "").strip()
        start_date = request.form.get("start_date", "").strip()
        end_date = request.form.get("end_date", "").strip()
        is_active = request.form.get("is_active") == "on"

        if action == "delete":
            if semester_id:
                Semester.query.filter_by(id=int(semester_id)).delete()
                db.session.commit()
                flash("Semester berhasil dihapus.", "info")
            return redirect(url_for("dashboard.staff_semesters"))

        if not name:
            flash("Nama semester wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_semesters"))

        if action == "update":
            semester = Semester.query.get(int(semester_id)) if semester_id else None
            if not semester:
                flash("Semester tidak ditemukan.", "error")
                return redirect(url_for("dashboard.staff_semesters"))
            if is_active:
                Semester.query.update({Semester.is_active: False})
            semester.name = name
            semester.start_date = start_date or None
            semester.end_date = end_date or None
            semester.is_active = is_active
            db.session.commit()
            flash("Semester berhasil diperbarui.", "info")
            return redirect(url_for("dashboard.staff_semesters"))

        if is_active:
            Semester.query.update({Semester.is_active: False})

        semester = Semester(
            name=name,
            start_date=start_date or None,
            end_date=end_date or None,
            is_active=is_active,
        )
        db.session.add(semester)
        db.session.commit()
        flash("Semester berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_semesters"))

    q = request.args.get("q", "").strip()
    query = Semester.query
    if q:
        query = query.filter(Semester.name.ilike(f"%{q}%"))
    semesters = query.order_by(Semester.id.desc()).limit(50).all()

    return render_template(
        "dashboard/staff_semesters.html",
        full_name=session.get("full_name"),
        semesters=semesters,
        q=q,
        active_semester=_get_active_semester(),
    )




@dashboard_bp.post("/staff/semester/select")
@login_required
@role_required("staff")
def staff_select_semester():
    semester_id = request.form.get("semester_id", "").strip()
    if semester_id:
        session["active_semester_id"] = int(semester_id)
    else:
        session.pop("active_semester_id", None)
    return redirect(request.referrer or url_for("dashboard.staff_home"))


# ---------------------------------------------------------------------------
# Holiday management
# ---------------------------------------------------------------------------

@dashboard_bp.route("/staff/holidays", methods=["GET", "POST"])
@login_required
@role_required("staff", "admin")
def staff_holidays():
    if request.method == "POST":
        action = request.form.get("action", "add")
        holiday_id = request.form.get("holiday_id", "").strip()

        if action == "delete" and holiday_id:
            Holiday.query.filter_by(id=int(holiday_id)).delete()
            db.session.commit()
            flash("Hari libur dihapus.", "info")
            return redirect(url_for("dashboard.staff_holidays"))

        date_str = request.form.get("date", "").strip()
        description = request.form.get("description", "").strip()
        if not date_str or not description:
            flash("Tanggal dan keterangan wajib diisi.", "error")
            return redirect(url_for("dashboard.staff_holidays"))

        try:
            import datetime as _dt
            parsed = _dt.date.fromisoformat(date_str)
        except ValueError:
            flash("Format tanggal tidak valid.", "error")
            return redirect(url_for("dashboard.staff_holidays"))

        existing = Holiday.query.filter_by(date=parsed).first()
        if existing:
            flash("Tanggal tersebut sudah terdaftar sebagai hari libur.", "error")
            return redirect(url_for("dashboard.staff_holidays"))

        db.session.add(Holiday(date=parsed, description=description))
        db.session.commit()
        flash("Hari libur berhasil ditambahkan.", "info")
        return redirect(url_for("dashboard.staff_holidays"))

    holidays = Holiday.query.order_by(Holiday.date.asc()).all()
    return render_template(
        "dashboard/staff_holidays.html",
        full_name=session.get("full_name"),
        holidays=holidays,
        active_semester=_get_active_semester(),
    )


# ---------------------------------------------------------------------------
# Pakta integritas template upload / download
# ---------------------------------------------------------------------------

@dashboard_bp.route("/staff/pakta-template", methods=["GET", "POST"])
@login_required
@role_required("staff", "admin")
def staff_pakta_template():
    if request.method == "POST":
        file = request.files.get("template_file")
        if not file or not file.filename:
            flash("File wajib dipilih.", "error")
            return redirect(url_for("dashboard.staff_pakta_template"))
        allowed_ext = {"pdf", "docx", "doc"}
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in allowed_ext:
            flash("Format file tidak didukung (PDF, DOCX, DOC).", "error")
            return redirect(url_for("dashboard.staff_pakta_template"))

        upload_root = current_app.config["UPLOAD_FOLDER"]
        dest_dir = os.path.join(upload_root, "templates")
        os.makedirs(dest_dir, exist_ok=True)
        original_name = secure_filename(file.filename)
        unique_name = f"{int(dt.datetime.utcnow().timestamp())}_{original_name}"
        file.save(os.path.join(dest_dir, unique_name))
        rel_path = os.path.join("templates", unique_name)

        record = PaktaTemplate(
            file_path=rel_path,
            original_filename=original_name,
            uploaded_by=session.get("user_id"),
        )
        db.session.add(record)
        db.session.commit()
        flash("Template pakta integritas berhasil diunggah.", "info")
        return redirect(url_for("dashboard.staff_pakta_template"))

    templates = PaktaTemplate.query.order_by(PaktaTemplate.id.desc()).limit(10).all()
    return render_template(
        "dashboard/staff_pakta_template.html",
        full_name=session.get("full_name"),
        templates=templates,
        active_semester=_get_active_semester(),
    )


@dashboard_bp.get("/staff/pakta-template/download/<int:template_id>")
@login_required
@role_required("staff", "admin")
def staff_download_pakta_template(template_id: int):
    tmpl = PaktaTemplate.query.get_or_404(template_id)
    upload_folder = current_app.config["UPLOAD_FOLDER"]
    dir_path = os.path.dirname(os.path.join(upload_folder, tmpl.file_path))
    filename = os.path.basename(tmpl.file_path)
    return send_from_directory(
        dir_path, filename,
        as_attachment=True,
        download_name=tmpl.original_filename or filename,
    )


# ---------------------------------------------------------------------------
# Excel upload for auto-fill schedules
# ---------------------------------------------------------------------------

@dashboard_bp.route("/staff/schedules/excel-template")
@login_required
@role_required("staff", "admin")
def staff_schedule_excel_template():
    """Generate and download an Excel template with dropdown validation from existing data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # Fetch existing data for dropdowns
    courses   = Course.query.order_by(Course.course_name).all()
    lecturers = Lecturer.query.order_by(Lecturer.lecturer_name).all()
    rooms     = Room.query.order_by(Room.room_name).all()
    buildings = Building.query.order_by(Building.building_name).all()

    course_names   = [c.course_name   for c in courses   if c.course_name]
    lecturer_names = [l.lecturer_name for l in lecturers if l.lecturer_name]
    room_names     = [r.room_name     for r in rooms     if r.room_name]
    building_names = [b.building_name for b in buildings if b.building_name]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Jadwal"

    # ── Hidden Referensi sheet for dropdown lists ──
    ref_ws = wb.create_sheet("Referensi")
    ref_ws.sheet_state = "hidden"
    for i, name in enumerate(course_names,   start=1): ref_ws.cell(row=i, column=1, value=name)
    for i, name in enumerate(lecturer_names, start=1): ref_ws.cell(row=i, column=2, value=name)
    for i, name in enumerate(room_names,     start=1): ref_ws.cell(row=i, column=3, value=name)
    for i, name in enumerate(building_names, start=1): ref_ws.cell(row=i, column=4, value=name)

    # ── Header row ──
    # New 10-column layout:
    # A: Matakuliah | B: Kelas | C: Pengajar 1 | D: Pengajar 2 | E: Pengajar 3
    # F: Ruangan | G: Hari | H: Jam Mulai | I: Jam Selesai | J: Gedung
    headers = [
        "Matakuliah", "Kelas",
        "Pengajar 1", "Pengajar 2 (opsional)", "Pengajar 3 (opsional)",
        "Ruangan", "Hari", "Jam Mulai", "Jam Selesai", "Gedung (opsional)",
    ]
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── Data Validation dropdowns (rows 2–500) ──
    MAX_ROW = 500

    if course_names:
        dv = DataValidation(type="list", formula1=f"Referensi!$A$1:$A${len(course_names)}", allow_blank=True, showDropDown=False)
        dv.sqref = f"A2:A{MAX_ROW}"
        ws.add_data_validation(dv)

    if lecturer_names:
        n = len(lecturer_names)
        for col_letter in ("C", "D", "E"):
            dv = DataValidation(type="list", formula1=f"Referensi!$B$1:$B${n}", allow_blank=True, showDropDown=False)
            dv.sqref = f"{col_letter}2:{col_letter}{MAX_ROW}"
            ws.add_data_validation(dv)

    if room_names:
        dv = DataValidation(type="list", formula1=f"Referensi!$C$1:$C${len(room_names)}", allow_blank=True, showDropDown=False)
        dv.sqref = f"F2:F{MAX_ROW}"
        ws.add_data_validation(dv)

    # Hari dropdown (hardcoded list) — column G
    dv_day = DataValidation(
        type="list",
        formula1='"Senin,Selasa,Rabu,Kamis,Jumat,Sabtu,Minggu"',
        allow_blank=True, showDropDown=False,
    )
    dv_day.sqref = f"G2:G{MAX_ROW}"
    ws.add_data_validation(dv_day)

    if building_names:
        dv = DataValidation(type="list", formula1=f"Referensi!$D$1:$D${len(building_names)}", allow_blank=True, showDropDown=False)
        dv.sqref = f"J2:J{MAX_ROW}"
        ws.add_data_validation(dv)

    # ── Note row (row 2) ──
    note_cell = ws.cell(
        row=2, column=1,
        value="Isi mulai baris 3. Kolom A/C/D/E/F/G/J memiliki dropdown. "
              "Kelas diisi A/B/C dst. Pengajar 2 & 3 opsional. Jam format HH:MM.",
    )
    note_cell.font = Font(italic=True, color="718096", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    note_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Column widths ──
    for i, w in enumerate([30, 7, 26, 26, 26, 20, 9, 10, 10, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="template_jadwal.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@dashboard_bp.route("/staff/schedules/excel-upload", methods=["GET", "POST"])
@login_required
@role_required("staff", "admin")
def staff_schedule_excel_upload():
    if request.method == "POST":
        file = request.files.get("excel_file")
        semester_id = request.form.get("semester_id", "").strip()
        if not file or not file.filename:
            flash("File Excel wajib dipilih.", "error")
            return redirect(url_for("dashboard.staff_schedule_excel_upload"))
        if not semester_id:
            flash("Semester wajib dipilih.", "error")
            return redirect(url_for("dashboard.staff_schedule_excel_upload"))
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in ("xlsx", "xls"):
            flash("Format file tidak didukung, gunakan .xlsx atau .xls.", "error")
            return redirect(url_for("dashboard.staff_schedule_excel_upload"))

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=3, values_only=True))
        except Exception as exc:
            flash(f"Gagal membaca file Excel: {exc}", "error")
            return redirect(url_for("dashboard.staff_schedule_excel_upload"))

        # Expected columns (1-indexed):
        # A: course_name/code, B: lecturer_name/nidn, C: room_name/code,
        # D: day_name, E: start_time (HH:MM), F: end_time (HH:MM)
        added = 0
        errors = []
        sem_id = int(semester_id)

        for i, row in enumerate(rows, start=2):
            if not any(row):
                continue
            try:
                # Column layout (10 cols):
                # A=Matakuliah, B=Kelas, C=Pengajar1, D=Pengajar2, E=Pengajar3,
                # F=Ruangan, G=Hari, H=JamMulai, I=JamSelesai, J=Gedung
                course_val    = str(row[0] or "").strip()
                class_name    = str(row[1] or "").strip()
                lecturer_val1 = str(row[2] or "").strip() if len(row) > 2 else ""
                lecturer_val2 = str(row[3] or "").strip() if len(row) > 3 else ""
                lecturer_val3 = str(row[4] or "").strip() if len(row) > 4 else ""
                room_val      = str(row[5] or "").strip() if len(row) > 5 else ""
                day_name      = str(row[6] or "").strip() if len(row) > 6 else ""
                start_str     = str(row[7] or "").strip() if len(row) > 7 else ""
                end_str       = str(row[8] or "").strip() if len(row) > 8 else ""
                building_val  = str(row[9] or "").strip() if len(row) > 9 else ""

                if not all([course_val, lecturer_val1, room_val, day_name, start_str, end_str]):
                    errors.append(f"Baris {i}: kolom wajib tidak lengkap (Matakuliah/Pengajar1/Ruangan/Hari/Jam), dilewati.")
                    continue

                # Resolve course
                course = (Course.query.filter(
                    or_(Course.course_name.ilike(course_val), Course.course_code.ilike(course_val))
                ).first())
                if not course:
                    course = Course(course_name=course_val)
                    db.session.add(course)
                    db.session.flush()

                # Resolve lecturers (Pengajar 1 wajib, 2 & 3 opsional)
                resolved_lecturers = []
                for lval in [lecturer_val1, lecturer_val2, lecturer_val3]:
                    if not lval:
                        continue
                    lec = Lecturer.query.filter(
                        or_(Lecturer.lecturer_name.ilike(lval), Lecturer.nidn.ilike(lval))
                    ).first()
                    if not lec:
                        lec = Lecturer(lecturer_name=lval)
                        db.session.add(lec)
                        db.session.flush()
                    resolved_lecturers.append(lec)

                # Resolve building (column J, opsional)
                building_obj = None
                if building_val:
                    building_obj = Building.query.filter(
                        Building.building_name.ilike(building_val)
                    ).first()
                    if not building_obj:
                        building_obj = Building(building_name=building_val)
                        db.session.add(building_obj)
                        db.session.flush()

                # Resolve room
                room = (Room.query.filter(
                    or_(Room.room_name.ilike(room_val), Room.room_code.ilike(room_val))
                ).first())
                if not room:
                    room = Room(
                        room_name=room_val,
                        room_type="Ruang Kelas",
                        building_id=building_obj.id if building_obj else None,
                    )
                    db.session.add(room)
                    db.session.flush()
                elif building_obj and not room.building_id:
                    room.building_id = building_obj.id

                # Parse time strings to datetime.time
                try:
                    start_time = dt.datetime.strptime(start_str, "%H:%M").time()
                    end_time   = dt.datetime.strptime(end_str,   "%H:%M").time()
                except ValueError:
                    errors.append(f"Baris {i}: format jam salah (gunakan HH:MM), dilewati.")
                    continue

                sched = ClassSchedule(
                    course_id=course.id,
                    room_id=room.id,
                    day_name=day_name,
                    start_time=start_time,
                    end_time=end_time,
                    semester_id=sem_id,
                    class_name=class_name,
                )
                sched.lecturers = resolved_lecturers
                db.session.add(sched)
                added += 1
            except Exception as exc:
                errors.append(f"Baris {i}: {exc}")

        db.session.commit()
        msg = f"{added} jadwal berhasil diimpor."
        if errors:
            msg += f" {len(errors)} baris gagal: " + "; ".join(errors[:5])
        flash(msg, "info" if added else "error")
        return redirect(url_for("dashboard.staff_schedules"))

    semesters = Semester.query.order_by(Semester.id.desc()).all()
    return render_template(
        "dashboard/staff_schedule_excel.html",
        full_name=session.get("full_name"),
        semesters=semesters,
        active_semester=_get_active_semester(),
        active_menu="schedule_excel",
    )


# ---------------------------------------------------------------------------
# QR code pages (staff generates QR for rooms)
# ---------------------------------------------------------------------------

@dashboard_bp.get("/staff/qr-codes")
@login_required
@role_required("staff", "admin")
def staff_qr_codes():
    rooms = Room.query.order_by(Room.room_name.asc()).all()
    return render_template(
        "dashboard/staff_qr_codes.html",
        full_name=session.get("full_name"),
        rooms=rooms,
        active_semester=_get_active_semester(),
    )

